import logging
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from models import *
from models.inv_session import InventoryStatus

logger = logging.getLogger(__name__)

async def create_inventory_report(session_id: int, db: AsyncSession) -> BytesIO:
    try:
        inv_session = await db.scalar(
            select(InventorySession)
            .options(
                joinedload(InventorySession.creator),
                joinedload(InventorySession.users).joinedload(InventoryUser.user),
                joinedload(InventorySession.products)
                .joinedload(InventoryProduct.product)
                .joinedload(Product.unit),
                joinedload(InventorySession.products)
                .joinedload(InventoryProduct.product)
                .joinedload(Product.category),
                joinedload(InventorySession.products).joinedload(InventoryProduct.user),
                joinedload(InventorySession.organization)
            )
            .where(InventorySession.id == session_id)
        )

        if not inv_session:
            raise ValueError("Session not found")

        wb = Workbook()
        ws = wb.active
        ws.title = "Инвентаризация"

        ws.merge_cells('A1:D1')
        ws['A1'] = f"Инвентаризационная ведомость №{inv_session.id} от {datetime.now().strftime('%d.%m.%Y г. %H:%M')}"
        ws['A1'].font = Font(bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        status_map = {
            InventoryStatus.finished: "Завершена",
            InventoryStatus.cancelled: "Отменена",
        }
        status_str = status_map.get(inv_session.status, "В процессе")
        ws['A3'] = f"Статус: {status_str}"

        ws['A5'] = f"Инициатор: {inv_session.creator.first_name} {inv_session.creator.last_name}"

        for i, user in enumerate(inv_session.users, start=1):
            ws[f'A{5 + i}'] = f"Участник {i}: {user.user.first_name} {user.user.last_name}"

        ws.append([])
        headers = ["Категория", "Продукт", "Ед.изм.", "Кол."]
        ws.append(headers)

        for cell in ws[8]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for inv_product in inv_session.products:
            product = inv_product.product
            ws.append([
                product.category.name if product.category else "",
                product.name,
                product.unit.name if product.unit else "ед.",
                inv_product.quantity
            ])

        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file

    except Exception as e:
        logger.error(f"Error creating inventory report: {e}", exc_info=True)
        raise
